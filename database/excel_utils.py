"""
Excel Import/Export Utilities for Asset Management System
Handles Excel file generation, template creation, validation, and import
"""

import pandas as pd
from io import BytesIO
from datetime import datetime
from typing import Tuple, Dict, List, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

# Import database functions
from database.db import create_asset, get_all_assets

# Column configuration for Excel
EXCEL_COLUMNS = [
    {"name": "Serial Number", "db_field": "serial_number", "required": True, "type": "text"},
    {"name": "Asset Type", "db_field": "asset_type", "required": False, "type": "dropdown",
     "options": ["Laptop", "Phone", "Printer", "Other"]},
    {"name": "Brand", "db_field": "brand", "required": False, "type": "dropdown",
     "options": ["Lenovo", "HP", "Dell", "Apple", "Asus", "Acer", "Other"]},
    {"name": "Model", "db_field": "model", "required": False, "type": "text"},
    {"name": "Specs", "db_field": "specs", "required": False, "type": "text"},
    {"name": "Touch Screen", "db_field": "touch_screen", "required": False, "type": "boolean"},
    {"name": "Processor", "db_field": "processor", "required": False, "type": "text"},
    {"name": "RAM (GB)", "db_field": "ram_gb", "required": False, "type": "integer"},
    {"name": "Storage Type", "db_field": "storage_type", "required": False, "type": "dropdown",
     "options": ["SSD", "HDD"]},
    {"name": "Storage (GB)", "db_field": "storage_gb", "required": False, "type": "integer"},
    {"name": "OS Installed", "db_field": "os_installed", "required": False, "type": "dropdown",
     "options": ["Windows 10 Pro", "Windows 11 Pro", "macOS", "Linux", "Chrome OS", "None"]},
    {"name": "Office License Key", "db_field": "office_license_key", "required": False, "type": "text"},
    {"name": "Device Password", "db_field": "device_password", "required": False, "type": "text"},
    {"name": "Current Status", "db_field": "current_status", "required": False, "type": "dropdown",
     "options": ["IN_STOCK_WORKING", "IN_STOCK_FAULTY", "WITH_CLIENT", "WITH_VENDOR_REPAIR",
                 "IN_OFFICE_TESTING", "RETURNED_FROM_CLIENT", "SOLD", "WRITTEN_OFF"]},
    {"name": "Current Location", "db_field": "current_location", "required": False, "type": "text"},
    {"name": "Purchase Date", "db_field": "purchase_date", "required": False, "type": "date"},
    {"name": "Purchase Price", "db_field": "purchase_price", "required": False, "type": "decimal"},
    {"name": "Notes", "db_field": "notes", "required": False, "type": "text"},
]

# Styling constants
HEADER_FILL = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_BORDER = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)


def export_assets_to_excel(df: pd.DataFrame) -> BytesIO:
    """
    Export assets DataFrame to a formatted Excel file.

    Args:
        df: DataFrame containing assets data

    Returns:
        BytesIO buffer containing the Excel file
    """
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Assets"

    # Define column order for export
    export_columns = [col["name"] for col in EXCEL_COLUMNS]

    # Map DataFrame columns to export columns
    column_mapping = {col["db_field"]: col["name"] for col in EXCEL_COLUMNS}

    # Rename columns in DataFrame
    df_export = df.copy()
    df_export = df_export.rename(columns=column_mapping)

    # Select only the columns we want to export (in order)
    available_columns = [col for col in export_columns if col in df_export.columns]
    df_export = df_export[available_columns]

    # Write header row
    for col_idx, col_name in enumerate(available_columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = CELL_BORDER

    # Write data rows
    for row_idx, row_data in enumerate(df_export.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="center")

    # Auto-adjust column widths
    for col_idx, col_name in enumerate(available_columns, 1):
        max_length = len(str(col_name))
        for row_idx in range(2, len(df_export) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def export_dataframe_to_excel(df: pd.DataFrame, sheet_name: str = "Data") -> BytesIO:
    """Export any DataFrame to formatted Excel with orange headers and borders."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    columns = [c for c in df.columns if not str(c).startswith("_")]
    df_export = df[columns].copy()

    # Header row
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = CELL_BORDER

    # Data rows
    for row_idx, row_data in enumerate(df_export.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            if pd.isna(value):
                value = None
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="center")

    # Auto-adjust column widths
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(str(col_name))
        for row_idx in range(2, min(len(df_export) + 2, 102)):  # Sample first 100 rows
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def detect_columns(df: pd.DataFrame) -> List[str]:
    """Return list of non-empty column names from uploaded DataFrame."""
    return [str(c).strip() for c in df.columns if str(c).strip()]


def auto_suggest_mapping(detected_cols: List[str]) -> Dict[str, str]:
    """
    Suggest an app field for each detected column using keyword matching.
    Returns {detected_col: suggested_app_field or '-- Skip --'}.
    First keyword match wins; unmatched columns default to '-- Skip --'.
    """
    KEYWORD_MAP = {
        "Serial Number":     ["serial", "sn", "s/n", "asset_id", "device_id", "asset id", "tag"],
        "Asset Type":        ["asset type", "device type", "type", "category"],
        "Brand":             ["brand", "make", "manufacturer", "vendor", "oem"],
        "Model":             ["model", "device name", "product name", "product"],
        "Specs":             ["specs", "specification", "config", "configuration"],
        "Processor":         ["processor", "cpu", "chip", "core"],
        "RAM (GB)":          ["ram", "memory", "mem"],
        "Storage (GB)":      ["storage gb", "storage size", "storage", "hdd", "ssd", "disk", "capacity"],
        "Storage Type":      ["storage type", "drive type", "disk type"],
        "OS Installed":      ["os installed", "operating system", "os", "windows", "macos", "linux"],
        "Touch Screen":      ["touch screen", "touchscreen", "touch"],
        "Purchase Date":     ["purchase date", "procurement date", "bought date", "buy date", "date purchased"],
        "Purchase Price":    ["purchase price", "price", "cost", "amount", "value", "rate"],
        "Current Status":    ["current status", "status", "condition", "state"],
        "Current Location":  ["current location", "location", "place", "site", "office", "assigned to"],
        "Notes":             ["notes", "remarks", "comments", "additional info"],
        "Office License Key":["office license", "license key", "product key", "office key", "license"],
        "Device Password":   ["device password", "password", "pwd", "pin"],
    }

    suggestions = {}
    used_fields = set()
    for col in detected_cols:
        col_lower = col.lower().strip()
        matched = "-- Skip --"
        for app_field, keywords in KEYWORD_MAP.items():
            if app_field in used_fields:
                continue
            if any(kw in col_lower for kw in keywords):
                matched = app_field
                break
        suggestions[col] = matched
        if matched != "-- Skip --":
            used_fields.add(matched)
    return suggestions


def apply_column_mapping(df: pd.DataFrame, mapping: Dict[str, str]) -> pd.DataFrame:
    """
    Rename/select columns based on user mapping.
    mapping = {'Their Column': 'App Field', 'Other Col': '-- Skip --'}
    Returns new DataFrame with only mapped (non-skipped) columns, renamed to app field names.
    """
    rename_map = {their_col: app_field
                  for their_col, app_field in mapping.items()
                  if app_field != "-- Skip --" and their_col in df.columns}
    if not rename_map:
        return pd.DataFrame()

    # Check for duplicate app field targets
    seen = {}
    for their_col, app_field in rename_map.items():
        if app_field in seen:
            # Keep first mapping, skip duplicates
            rename_map[their_col] = "-- Skip --"
        else:
            seen[app_field] = their_col

    # Remove any newly skipped
    rename_map = {k: v for k, v in rename_map.items() if v != "-- Skip --"}
    if not rename_map:
        return pd.DataFrame()

    return df[list(rename_map.keys())].rename(columns=rename_map).copy()


def generate_import_template() -> BytesIO:
    """
    Generate a blank Excel template for importing assets.
    Includes headers, data validation dropdowns, and instructions.

    Returns:
        BytesIO buffer containing the template Excel file
    """
    wb = Workbook()

    # Main data sheet
    ws = wb.active
    ws.title = "Assets"

    # Write headers with styling
    for col_idx, col_config in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_config["name"])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = CELL_BORDER

        # Mark required columns
        if col_config["required"]:
            cell.value = f"{col_config['name']} *"

    # Add data validation for dropdown columns
    for col_idx, col_config in enumerate(EXCEL_COLUMNS, 1):
        if col_config["type"] == "dropdown" and "options" in col_config:
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            options_str = ",".join(col_config["options"])
            dv = DataValidation(
                type="list",
                formula1=f'"{options_str}"',
                allow_blank=True,
                showDropDown=False
            )
            dv.error = f"Please select from: {options_str}"
            dv.errorTitle = "Invalid Selection"
            dv.prompt = f"Select {col_config['name']}"
            dv.promptTitle = col_config["name"]
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}1000")

        elif col_config["type"] == "boolean":
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            dv = DataValidation(
                type="list",
                formula1='"TRUE,FALSE"',
                allow_blank=True
            )
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}1000")

    # Set column widths
    column_widths = {
        "Serial Number": 20,
        "Asset Type": 12,
        "Brand": 12,
        "Model": 20,
        "Specs": 30,
        "Touch Screen": 12,
        "Processor": 25,
        "RAM (GB)": 10,
        "Storage Type": 12,
        "Storage (GB)": 12,
        "OS Installed": 18,
        "Office License Key": 25,
        "Device Password": 15,
        "Current Status": 20,
        "Current Location": 25,
        "Purchase Date": 15,
        "Purchase Price": 15,
        "Notes": 35,
    }

    for col_idx, col_config in enumerate(EXCEL_COLUMNS, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        width = column_widths.get(col_config["name"], 15)
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add sample row (row 2) as example
    sample_data = {
        "Serial Number": "SAMPLE-001",
        "Asset Type": "Laptop",
        "Brand": "Lenovo",
        "Model": "ThinkPad X1 Carbon",
        "Specs": "14-inch FHD Display",
        "Touch Screen": "FALSE",
        "Processor": "Intel Core i7-1260P",
        "RAM (GB)": 16,
        "Storage Type": "SSD",
        "Storage (GB)": 512,
        "OS Installed": "Windows 11 Pro",
        "Office License Key": "",
        "Device Password": "",
        "Current Status": "IN_STOCK_WORKING",
        "Current Location": "Office",
        "Purchase Date": datetime.now().strftime("%Y-%m-%d"),
        "Purchase Price": 1500.00,
        "Notes": "Sample entry - delete this row before import",
    }

    for col_idx, col_config in enumerate(EXCEL_COLUMNS, 1):
        value = sample_data.get(col_config["name"], "")
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.font = Font(italic=True, color="888888")

    # Create Instructions sheet
    instructions_ws = wb.create_sheet("Instructions")
    instructions = [
        ["Asset Import Template Instructions"],
        [""],
        ["1. Fill in the 'Assets' sheet with your asset data"],
        ["2. Serial Number is REQUIRED - each asset must have a unique serial number"],
        ["3. Columns marked with * are required"],
        ["4. Use dropdown menus where available for consistent data"],
        ["5. Date format: YYYY-MM-DD (e.g., 2024-01-15)"],
        ["6. Touch Screen: Use TRUE or FALSE"],
        ["7. Delete the sample row (row 2) before importing"],
        ["8. Do not modify column headers"],
        [""],
        ["Valid Status Values:"],
        ["- IN_STOCK_WORKING: Asset is in stock and working"],
        ["- IN_STOCK_FAULTY: Asset is in stock but has issues"],
        ["- WITH_CLIENT: Asset is deployed to a client"],
        ["- WITH_VENDOR_REPAIR: Asset is with vendor for repair"],
        ["- IN_OFFICE_TESTING: Asset is being tested in office"],
        ["- RETURNED_FROM_CLIENT: Asset returned from client"],
        ["- SOLD: Asset has been sold"],
        ["- WRITTEN_OFF: Asset has been written off"],
    ]

    for row_idx, row_data in enumerate(instructions, 1):
        cell = instructions_ws.cell(row=row_idx, column=1, value=row_data[0] if row_data else "")
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif row_data and row_data[0].startswith("-"):
            cell.font = Font(italic=True)

    instructions_ws.column_dimensions["A"].width = 60

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def validate_import_data(df: pd.DataFrame) -> Tuple[bool, List[Dict], List[Dict], pd.DataFrame]:
    """
    Validate imported Excel data before inserting into database.

    Args:
        df: DataFrame from uploaded Excel file

    Returns:
        Tuple of (is_valid, errors, warnings, cleaned_df)
    """
    errors = []
    warnings = []

    # Check if DataFrame is empty
    if df.empty:
        errors.append({"row": 0, "field": "File", "message": "The uploaded file is empty"})
        return False, errors, warnings, df

    # Step 1: Normalize column names (strip whitespace, remove asterisks)
    df = df.copy()
    df.columns = [str(col).replace(" *", "").strip() for col in df.columns]

    # Step 2: Check required columns exist
    required_columns = [col["name"] for col in EXCEL_COLUMNS if col["required"]]
    missing_columns = [col for col in required_columns if col not in df.columns]

    if missing_columns:
        errors.append({
            "row": 0,
            "field": "Columns",
            "message": f"Missing required columns: {', '.join(missing_columns)}"
        })
        return False, errors, warnings, df

    # Get existing serial numbers from database
    existing_assets = get_all_assets()
    existing_serials = set()
    if existing_assets is not None and not existing_assets.empty:
        if "Serial Number" in existing_assets.columns:
            existing_serials = set(existing_assets["Serial Number"].dropna().astype(str).str.strip().str.upper())
        elif "serial_number" in existing_assets.columns:
            existing_serials = set(existing_assets["serial_number"].dropna().astype(str).str.strip().str.upper())

    # Track serial numbers in this import for duplicate detection
    import_serials = set()
    valid_rows = []
    empty_row_count = 0

    # Validate each row
    for idx, row in df.iterrows():
        row_num = idx + 2  # Excel row number (1-indexed + header)
        row_errors = []

        # Step 3: Skip completely empty rows early
        if row.isna().all():
            empty_row_count += 1
            continue

        # Check if all values are empty strings or NaN
        row_str_values = row.astype(str).str.strip()
        if (row_str_values == "").all() or (row_str_values == "nan").all():
            empty_row_count += 1
            continue

        # Step 4: Get and validate Serial Number FIRST
        serial_raw = row.get("Serial Number", "")
        serial = str(serial_raw).strip() if pd.notna(serial_raw) else ""

        # Handle NaN string
        if serial.lower() == "nan":
            serial = ""

        # Skip sample row
        if serial.upper() == "SAMPLE-001":
            warnings.append({
                "row": row_num,
                "field": "Serial Number",
                "message": "Sample row skipped"
            })
            continue

        # Check Serial Number (required)
        if not serial:
            row_errors.append({
                "row": row_num,
                "field": "Serial Number",
                "message": "Serial Number is required"
            })
        else:
            # Check for duplicates in database
            if serial.upper() in existing_serials:
                warnings.append({
                    "row": row_num,
                    "field": "Serial Number",
                    "message": f"Serial Number '{serial}' already exists in database - will be skipped"
                })
                continue

            # Check for duplicates in this import
            if serial.upper() in import_serials:
                row_errors.append({
                    "row": row_num,
                    "field": "Serial Number",
                    "message": f"Duplicate Serial Number '{serial}' in import file"
                })
            else:
                import_serials.add(serial.upper())

        # Step 5: Validate numeric fields
        for field in ["RAM (GB)", "Storage (GB)"]:
            if field in df.columns:
                value = row.get(field)
                if pd.notna(value):
                    value_str = str(value).strip()
                    if value_str and value_str.lower() != "nan":
                        try:
                            int(float(value_str))
                        except (ValueError, TypeError):
                            row_errors.append({
                                "row": row_num,
                                "field": field,
                                "message": f"Invalid number: '{value}'"
                            })

        # Validate Purchase Price
        if "Purchase Price" in df.columns:
            value = row.get("Purchase Price")
            if pd.notna(value):
                value_str = str(value).strip()
                if value_str and value_str.lower() != "nan":
                    try:
                        float(value_str.replace(",", ""))
                    except (ValueError, TypeError):
                        row_errors.append({
                            "row": row_num,
                            "field": "Purchase Price",
                            "message": f"Invalid price: '{value}'"
                        })

        # Step 6: Validate dropdown values
        for col_config in EXCEL_COLUMNS:
            if col_config["type"] == "dropdown" and col_config["name"] in df.columns:
                value = row.get(col_config["name"])
                if pd.notna(value):
                    value_str = str(value).strip()
                    if value_str and value_str.lower() != "nan":
                        if value_str not in col_config["options"]:
                            warnings.append({
                                "row": row_num,
                                "field": col_config["name"],
                                "message": f"Value '{value_str}' not in standard options: {', '.join(col_config['options'][:3])}..."
                            })

        if row_errors:
            errors.extend(row_errors)
        else:
            valid_rows.append(idx)

    # Log empty row count as info (not error)
    if empty_row_count > 0:
        warnings.append({
            "row": 0,
            "field": "Info",
            "message": f"{empty_row_count} empty row(s) were skipped"
        })

    # Filter DataFrame to only valid rows
    if valid_rows:
        cleaned_df = df.loc[valid_rows].copy()
    else:
        cleaned_df = pd.DataFrame()

    is_valid = len(errors) == 0 and not cleaned_df.empty

    return is_valid, errors, warnings, cleaned_df


def import_assets_from_dataframe(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Import assets from a validated DataFrame into the database.

    Args:
        df: Cleaned and validated DataFrame

    Returns:
        Dictionary with import results
    """
    results = {
        "success": 0,
        "failed": 0,
        "errors": [],
        "imported_serials": []
    }

    # Normalize column names (strip whitespace, remove asterisks)
    df = df.copy()
    df.columns = [str(col).replace(" *", "").strip() for col in df.columns]

    for idx, row in df.iterrows():
        row_num = idx + 2  # Excel row number (1-indexed + header)

        try:
            # Skip completely empty rows
            if row.isna().all():
                continue

            # Convert row to string and check if all empty
            row_str_values = row.astype(str).str.strip()
            if (row_str_values == "").all() or (row_str_values == "nan").all():
                continue

            # Get serial number first for error reporting
            serial_raw = row.get("Serial Number", "")
            serial = str(serial_raw).strip() if pd.notna(serial_raw) else ""

            # Skip rows with no serial number
            if not serial or serial.lower() == "nan":
                results["failed"] += 1
                results["errors"].append({
                    "serial": f"Row {row_num}",
                    "error": "Missing required field: Serial Number"
                })
                continue

            # Skip sample row
            if serial.upper() == "SAMPLE-001":
                continue

            # Build asset data dictionary using EXCEL COLUMN NAMES (not db_field)
            # This is required because create_asset() expects Airtable-style field names
            asset_data = {}
            missing_fields = []

            for col_config in EXCEL_COLUMNS:
                col_name = col_config["name"]

                if col_name not in df.columns:
                    continue

                value = row.get(col_name)

                # Handle NaN/None/empty values
                if pd.isna(value):
                    value = None
                elif isinstance(value, str):
                    value = value.strip()
                    if value == "" or value.lower() == "nan":
                        value = None

                # Skip None values
                if value is None:
                    continue

                # Convert value based on type
                try:
                    if col_config["type"] == "integer":
                        value = int(float(str(value)))
                    elif col_config["type"] == "decimal":
                        value = float(str(value).replace(",", ""))
                    elif col_config["type"] == "boolean":
                        value = str(value).upper() in ("TRUE", "YES", "1")
                    elif col_config["type"] == "date":
                        if isinstance(value, str):
                            value = datetime.strptime(value, "%Y-%m-%d").date()
                        # datetime objects are passed through
                    else:
                        value = str(value).strip()
                except (ValueError, TypeError) as e:
                    # Skip fields with conversion errors
                    continue

                # Only add non-empty values
                if value is not None and value != "":
                    # Use the COLUMN NAME as key (not db_field) because
                    # create_asset() expects Airtable-style field names
                    asset_data[col_name] = value

            # Validate we have required data
            if "Serial Number" not in asset_data:
                results["failed"] += 1
                results["errors"].append({
                    "serial": serial or f"Row {row_num}",
                    "error": "Missing required field: Serial Number"
                })
                continue

            # Check if we have any meaningful data beyond serial number
            if len(asset_data) < 1:
                results["failed"] += 1
                results["errors"].append({
                    "serial": serial,
                    "error": "Row contains no valid data"
                })
                continue

            # Create asset in database
            success, asset_id, error = create_asset(asset_data)

            if success:
                results["success"] += 1
                results["imported_serials"].append(serial)
            else:
                results["failed"] += 1
                # Provide more specific error messages
                error_msg = error or "Unknown error"
                if "No data provided" in str(error_msg):
                    error_msg = "Row has no valid field values after processing"
                results["errors"].append({
                    "serial": serial,
                    "error": error_msg
                })

        except Exception as e:
            serial = str(row.get("Serial Number", f"Row {row_num}"))
            results["failed"] += 1
            results["errors"].append({
                "serial": serial,
                "error": f"Processing error: {str(e)}"
            })

    return results
